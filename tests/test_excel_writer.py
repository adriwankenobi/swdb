from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from scripts.excel_writer import update_excel


# Layout: YEAR, MEDIUM, SERIES, SERIES #, TITLE, #, AUTHOR, PUBLISHER,
# RELEASE, INFO, COVER[, ID]. AUTHOR is index 6, COVER index 10, ID index 11.
HEADER = [
    "YEAR", "MEDIUM", "SERIES", "SERIES #", "TITLE", "#",
    "AUTHOR", "PUBLISHER", "RELEASE", "INFO", "COVER",
]
HEADER_WITH_ID = [*HEADER, "ID"]


@pytest.fixture
def tiny_xlsx(tmp_path: Path) -> Path:
    """Create a minimal xlsx with one sheet matching an ERA_INDEX name."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REBELLION"
    ws.append(HEADER)
    ws.append([
        "0 ABY", "Novel", "Star Wars Episode", "IV", "A New Hope", "IV",
        "OLD AUTHOR", "OLD PUBLISHER", "1976.01.01", None, "OLD COVER",
    ])
    ws.append([
        "1 ABY", "Comic", None, None, "Some Comic", "1",
        "ANOTHER AUTHOR", None, None, None, None,
    ])
    path = tmp_path / "test.xlsx"
    wb.save(path)
    wb.close()
    return path


def test_update_excel_fills_empty_cells(tmp_path: Path):
    """Writer fills empty cells with parser values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REBELLION"
    ws.append(HEADER)
    ws.append([
        "0 ABY", "Novel", "Star Wars Episode", "IV", "A New Hope", "IV",
        None, None, None, None, None,  # all enriched cells empty
    ])
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    wb.close()

    enriched = {
        (5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): {
            "authors": ["Alan Dean Foster"],
            "publisher": "Ballantine Books",
            "release_date": "1976-11-12",
            "release_precision": "day",
            "cover_url": "https://example.com/cover.jpg",
        },
    }
    result = update_excel(path, enriched)
    assert result["updated"] == 1

    wb = load_workbook(path, data_only=True)
    ws = wb["REBELLION"]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row2[6] == "Alan Dean Foster"
    assert row2[7] == "Ballantine Books"
    assert row2[8] == "1976.11.12"
    assert row2[10] == "https://example.com/cover.jpg"
    wb.close()


def test_update_excel_does_not_overwrite_populated_cells(tiny_xlsx: Path):
    """Cells that already have a value are NEVER overwritten by the parser."""
    enriched = {
        (5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): {
            "authors": ["Alan Dean Foster"],
            "publisher": "Ballantine Books",
            "release_date": "1976-11-12",
            "release_precision": "day",
            "cover_url": "https://example.com/cover.jpg",
        },
    }
    result = update_excel(tiny_xlsx, enriched)
    # tiny_xlsx row 2 is fully populated with OLD values; nothing changes.
    assert result["updated"] == 0

    wb = load_workbook(tiny_xlsx, data_only=True)
    ws = wb["REBELLION"]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row2[6] == "OLD AUTHOR"
    assert row2[7] == "OLD PUBLISHER"
    assert row2[8] == "1976.01.01"
    assert row2[10] == "OLD COVER"
    wb.close()


@pytest.mark.parametrize(
    "precision,expected",
    [("day", "1976.11.12"), ("month", "1976.11"), ("year", "1976")],
)
def test_update_excel_writes_release_at_precision(tmp_path: Path, precision, expected):
    wb = Workbook()
    ws = wb.active
    ws.title = "REBELLION"
    ws.append(HEADER)
    ws.append([
        "0 ABY", "Novel", "Star Wars Episode", "IV", "A New Hope", "IV",
        None, None, None, None, None,
    ])
    path = tmp_path / f"prec-{precision}.xlsx"
    wb.save(path)
    wb.close()

    enriched = {
        (5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): {
            "release_date": "1976-11-12",
            "release_precision": precision,
        },
    }
    update_excel(path, enriched)
    wb = load_workbook(path, data_only=True)
    ws = wb["REBELLION"]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row2[8] == expected
    wb.close()


def test_update_excel_does_not_touch_unrelated_rows(tiny_xlsx: Path):
    enriched = {
        (5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): {
            "authors": ["Alan Dean Foster"],
        },
    }
    update_excel(tiny_xlsx, enriched)
    wb = load_workbook(tiny_xlsx, data_only=True)
    ws = wb["REBELLION"]
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    # Some Comic row was not in lookup — author cell stays as-is
    assert row3[6] == "ANOTHER AUTHOR"
    wb.close()


def test_update_excel_skips_missing_fields(tmp_path: Path):
    """Fields not in enriched dict do not affect populated cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REBELLION"
    ws.append(HEADER)
    ws.append([
        "0 ABY", "Novel", "Star Wars Episode", "IV", "A New Hope", "IV",
        None, "OLD PUBLISHER", "1976.01.01", None, "OLD COVER",
    ])
    path = tmp_path / "partial.xlsx"
    wb.save(path)
    wb.close()

    enriched = {
        (5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): {
            "authors": ["Alan Dean Foster"],  # only authors provided
        },
    }
    update_excel(path, enriched)
    wb = load_workbook(path, data_only=True)
    ws = wb["REBELLION"]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row2[6] == "Alan Dean Foster"  # was empty, now filled
    assert row2[7] == "OLD PUBLISHER"     # untouched
    assert row2[8] == "1976.01.01"        # untouched
    assert row2[10] == "OLD COVER"        # untouched
    wb.close()


def test_update_excel_stamps_id_header_when_missing(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "REBELLION"
    ws.append(HEADER)  # no ID header
    ws.append([
        "0 ABY", "Novel", "Star Wars Episode", "IV", "A New Hope", "IV",
        None, None, None, None, None,
    ])
    path = tmp_path / "noheader.xlsx"
    wb.save(path)
    wb.close()

    ids = {(5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): "gen-001"}
    update_excel(path, {}, ids=ids)

    wb = load_workbook(path, data_only=True)
    ws = wb["REBELLION"]
    assert ws.cell(row=1, column=12).value == "ID"  # ID is column L (12)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][11] == "gen-001"
    wb.close()


def test_update_excel_writes_generated_ids(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "REBELLION"
    ws.append(HEADER_WITH_ID)
    ws.append([
        "0 ABY", "Novel", "Star Wars Episode", "IV", "A New Hope", "IV",
        None, None, None, None, None, None,  # blank ID
    ])
    ws.append([
        "1 ABY", "Comic", None, None, "Some Comic", "1",
        None, None, None, None, None, "already-set",  # ID present
    ])
    path = tmp_path / "ids.xlsx"
    wb.save(path)
    wb.close()

    ids = {
        (5, "A New Hope", "Star Wars Episode", "Novel", "IV", "IV"): "gen-001",
        (5, "Some Comic", None, "Comic", None, "1"): "should-not-overwrite",
    }
    update_excel(path, {}, ids=ids)

    wb = load_workbook(path, data_only=True)
    ws = wb["REBELLION"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][11] == "gen-001"          # blank cell filled
    assert rows[1][11] == "already-set"       # present cell untouched
    wb.close()


def test_update_excel_keeps_issues_of_a_series_separate(tmp_path: Path):
    """Rows differing only by the per-work "#" must not receive each other's
    data. Before the "#" joined the lookup key they collapsed onto one entry
    and every issue got the last one's author/release/cover."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NON-CANON"
    ws.append(HEADER_WITH_ID)
    for n in (1, 2):
        ws.append([
            None, "Comic", "Infinities", None, "A New Hope", n,
            None, None, None, None, None, None,
        ])
    path = tmp_path / "issues.xlsx"
    wb.save(path)
    wb.close()

    enriched = {
        (9, "A New Hope", "Infinities", "Comic", None, "1"): {
            "authors": ["Chris Warner", "Drew Johnson"],
            "release_date": "2001-05-02",
            "release_precision": "day",
            "cover_url": "https://example.com/anh1.jpg",
        },
        (9, "A New Hope", "Infinities", "Comic", None, "2"): {
            "authors": ["Chris Warner", "Al Rio"],
            "release_date": "2001-06-06",
            "release_precision": "day",
            "cover_url": "https://example.com/anh2.jpg",
        },
    }
    ids = {
        (9, "A New Hope", "Infinities", "Comic", None, "1"): "id-issue-1",
        (9, "A New Hope", "Infinities", "Comic", None, "2"): "id-issue-2",
    }
    result = update_excel(path, enriched, ids=ids)
    assert result["updated"] == 2

    wb = load_workbook(path, data_only=True)
    rows = list(wb["NON-CANON"].iter_rows(min_row=2, values_only=True))
    assert rows[0][6] == "Chris Warner, Drew Johnson"
    assert rows[0][8] == "2001.05.02"
    assert rows[0][10] == "https://example.com/anh1.jpg"
    assert rows[0][11] == "id-issue-1"
    assert rows[1][6] == "Chris Warner, Al Rio"
    assert rows[1][8] == "2001.06.06"
    assert rows[1][10] == "https://example.com/anh2.jpg"
    assert rows[1][11] == "id-issue-2"
    wb.close()
