from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.excel_colors import ColorResolver, apply_tint, parse_theme_colors

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "Star Wars EU.xlsx"


THEME_XML_SAMPLE = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Sheets">
  <a:themeElements>
    <a:clrScheme name="Sheets">
      <a:dk1><a:srgbClr val="000000"/></a:dk1>
      <a:lt1><a:srgbClr val="FFFFFF"/></a:lt1>
      <a:dk2><a:srgbClr val="111111"/></a:dk2>
      <a:lt2><a:srgbClr val="EEEEEE"/></a:lt2>
      <a:accent1><a:srgbClr val="4285F4"/></a:accent1>
      <a:accent2><a:srgbClr val="EA4335"/></a:accent2>
      <a:accent3><a:srgbClr val="FBBC04"/></a:accent3>
      <a:accent4><a:srgbClr val="34A853"/></a:accent4>
      <a:accent5><a:srgbClr val="FF6D01"/></a:accent5>
      <a:accent6><a:srgbClr val="46BDC6"/></a:accent6>
      <a:hlink><a:srgbClr val="1155CC"/></a:hlink>
      <a:folHlink><a:srgbClr val="996633"/></a:folHlink>
    </a:clrScheme>
  </a:themeElements>
</a:theme>"""


def test_apply_tint_zero_returns_input():
    assert apply_tint("FF8800", 0.0) == "FF8800"


def test_apply_tint_positive_lightens():
    out = apply_tint("FF0000", 0.5)
    r, g, b = int(out[0:2], 16), int(out[2:4], 16), int(out[4:6], 16)
    assert r == 255
    assert g > 0
    assert b > 0


def test_apply_tint_negative_darkens():
    out = apply_tint("808080", -0.5)
    r = int(out[0:2], 16)
    assert 60 <= r <= 70


def test_apply_tint_pinned_workbook_combos():
    # Regression pins for the three theme+tint pairs the workbook actually uses.
    # Computed once via Microsoft's HLS-tint formula; lock them down so future
    # refactors of the formula are caught.
    assert apply_tint("FBBC04", 0.8) == "FEF2CD"  # accent3 yellow, light tint
    assert apply_tint("EA4335", 0.6) == "F7B4AE"  # accent2 red, mid tint
    assert apply_tint("46BDC6", 0.8) == "DAF2F4"  # accent6 teal, light tint


def test_parse_theme_colors_returns_12_colors_with_lt_dk_swap():
    colors = parse_theme_colors(THEME_XML_SAMPLE)
    assert len(colors) == 12
    # openpyxl swaps lt1<->dk1 and lt2<->dk2 vs XML order.
    assert colors[0] == "FFFFFF"  # lt1
    assert colors[1] == "000000"  # dk1
    assert colors[2] == "EEEEEE"  # lt2
    assert colors[3] == "111111"  # dk2
    assert colors[4] == "4285F4"  # accent1
    assert colors[5] == "EA4335"  # accent2
    assert colors[6] == "FBBC04"  # accent3
    assert colors[9] == "46BDC6"  # accent6


def test_parse_theme_colors_handles_sysclr():
    # <a:sysClr val="windowText" lastClr="000000"/> should resolve to lastClr.
    xml = b"""<?xml version="1.0"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <a:themeElements><a:clrScheme name="x">
    <a:dk1><a:sysClr val="windowText" lastClr="222222"/></a:dk1>
    <a:lt1><a:sysClr val="window" lastClr="DDDDDD"/></a:lt1>
    <a:dk2><a:srgbClr val="000000"/></a:dk2>
    <a:lt2><a:srgbClr val="FFFFFF"/></a:lt2>
    <a:accent1><a:srgbClr val="111111"/></a:accent1>
    <a:accent2><a:srgbClr val="222222"/></a:accent2>
    <a:accent3><a:srgbClr val="333333"/></a:accent3>
    <a:accent4><a:srgbClr val="444444"/></a:accent4>
    <a:accent5><a:srgbClr val="555555"/></a:accent5>
    <a:accent6><a:srgbClr val="666666"/></a:accent6>
    <a:hlink><a:srgbClr val="777777"/></a:hlink>
    <a:folHlink><a:srgbClr val="888888"/></a:folHlink>
  </a:clrScheme></a:themeElements></a:theme>"""
    colors = parse_theme_colors(xml)
    # After lt/dk swap: index 0 = lt1 (DDDDDD), index 1 = dk1 (222222)
    assert colors[0] == "DDDDDD"
    assert colors[1] == "222222"


@pytest.fixture(scope="module")
def workbook():
    return load_workbook(EXCEL_PATH, data_only=True)


def test_resolver_returns_hex_for_filled_row(workbook):
    resolver = ColorResolver(workbook)
    sheet = workbook["REBELLION"]
    cell = sheet.cell(row=2, column=1)
    color = resolver.resolve(cell)
    assert color is not None
    assert color.startswith("#") and len(color) == 7


def test_resolver_returns_none_for_no_fill_row(workbook):
    resolver = ColorResolver(workbook)
    sheet = workbook["REBELLION"]
    no_fill_cell = None
    for row in sheet.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.fill.patternType is None and cell.value is None:
            no_fill_cell = cell
            break
    assert no_fill_cell is not None, "expected a no-fill row in REBELLION"
    assert resolver.resolve(no_fill_cell) is None


def test_resolver_handles_raw_rgb():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="x")
    cell.fill = PatternFill(start_color="FF112233", end_color="FF112233", fill_type="solid")
    resolver = ColorResolver(wb)
    assert resolver.resolve(cell) == "#112233"
