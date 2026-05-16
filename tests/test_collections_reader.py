"""Tests for the COLLECTIONS-sheet reader."""

from openpyxl import Workbook

from scripts.collections_reader import ExcelCollectionRow, read_collections


def _build(path):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("COLLECTIONS")
    ws.append(["TITLE", "RELEASE", "INFO", "COVER"])
    ws.append(["Dark Empire (TPB)", "1995.01.01", "https://example/wiki/Dark_Empire_TPB", None])
    ws.append(["Empty Row", None, None, None])
    ws.append([None, None, None, None])  # truly empty
    wb.save(path)


def test_read_collections_yields_non_empty_rows(tmp_path):
    path = tmp_path / "x.xlsx"
    _build(path)
    rows = list(read_collections(path))
    assert len(rows) == 2
    assert rows[0] == ExcelCollectionRow(
        title="Dark Empire (TPB)",
        release_date_str="1995.01.01",
        info_url="https://example/wiki/Dark_Empire_TPB",
        cover_url=None,
        color=None,
    )
    assert rows[1].title == "Empty Row"
    assert rows[1].release_date_str is None


def test_read_collections_returns_empty_when_sheet_missing(tmp_path):
    wb = Workbook()
    path = tmp_path / "no_sheet.xlsx"
    wb.save(path)
    assert list(read_collections(path)) == []
