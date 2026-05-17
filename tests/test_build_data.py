"""Tests for the build orchestrator's row-to-work conversion."""

from unittest.mock import MagicMock

from scripts import build_data
from scripts.build_data import _enrich, _row_to_work, _split_series_and_number
from scripts.excel_reader import ExcelRow


def _row(**over) -> ExcelRow:
    base = dict(
        era=1,
        title="Knight Errant",
        series=None,
        medium="Novel",
        number=None,
        year=-1032,
        year_end=None,
        info_url=None,
        cover_url=None,
        color=None,
        author=None,
        publisher=None,
        release_date_str=None,
        collected=None,
    )
    base.update(over)
    return ExcelRow(**base)


def test_row_to_work_emits_string_era():
    work = _row_to_work(_row(era=1))
    assert work["era"] == "OLD REPUBLIC"


def test_row_to_work_emits_string_medium():
    work = _row_to_work(_row(medium="Novel"))
    assert work["medium"] == "Novel"


def test_row_to_work_id_is_stable_across_schema_change():
    # ID must not change when we switch the JSON shape; make_id consumes
    # `int` era and the canonical medium STRING, so the canonical key string
    # is identical to the pre-refactor era=int / medium=str pipeline.
    work = _row_to_work(
        _row(
            era=5,
            title="A New Hope",
            series="Star Wars Episode",
            medium="Novel",
            number="IV",
            year=0,
        )
    )
    # Frozen value captured from main before any code changes via:
    #   make_id(era=5, series='Star Wars Episode', title='A New Hope',
    #           medium='Novel', number='IV')
    assert work["id"] == "34a13f75-b121-5c91-b435-f765f951e4a5"


# ---------------------------------------------------------------------------
# _enrich tests
# ---------------------------------------------------------------------------


def _full_excel_row(**over):
    return _row(
        info_url="https://example.com/wiki/page",
        cover_url="https://example.com/cover.jpg",
        author="Alan Dean Foster",
        publisher="Ballantine Books",
        release_date_str="1976.11.12",
        **over,
    )


def test_enrich_full_excel_skips_fetch_html():
    row = _full_excel_row()
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = (row.info_url, "from_excel")
    client.verify_url_alive.return_value = True
    unmatched, dead = [], []

    _enrich(work, row, client, unmatched, dead)

    client.fetch_html.assert_not_called()
    assert work["wiki_url"] == row.info_url
    assert work["authors"] == ["Alan Dean Foster"]
    assert work["publisher"] == "Ballantine Books"
    assert work["release_date"] == "1976-11-12"
    assert work["release_precision"] == "day"
    assert work["cover_url"] == row.cover_url
    assert dead == []


def test_enrich_full_excel_logs_dead_wiki_url():
    row = _full_excel_row()
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = (row.info_url, "from_excel")
    client.verify_url_alive.side_effect = lambda u: u != row.info_url
    unmatched, dead = [], []

    _enrich(work, row, client, unmatched, dead)

    assert any("wiki" in entry and row.info_url in entry for entry in dead)


def test_enrich_full_excel_logs_dead_cover_url():
    row = _full_excel_row()
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = (row.info_url, "from_excel")
    client.verify_url_alive.side_effect = lambda u: u != row.cover_url
    unmatched, dead = [], []

    _enrich(work, row, client, unmatched, dead)

    assert any("cover" in entry and row.cover_url in entry for entry in dead)


def test_enrich_partial_excel_fetches_and_excel_wins_wholesale(monkeypatch):
    # Excel has author + cover; missing publisher + release_date.
    row = _row(
        info_url="https://example.com/wiki/page",
        cover_url="https://example.com/cover.jpg",
        author="Excel Author",
        publisher=None,
        release_date_str=None,
    )
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = (row.info_url, "from_excel")
    client.fetch_html.return_value = "<html>infobox</html>"

    def fake_parse(html):
        return {
            "authors": ["Parser Author A", "Parser Author B"],  # discarded
            "publisher": "Parser Publisher",  # used
            "release_date": "1980-05-20",  # used
            "release_precision": "day",
            "cover_url": "https://example.com/parser-cover.jpg",  # discarded
        }

    monkeypatch.setattr("scripts.build_data.parse_infobox", fake_parse)
    unmatched, dead = [], []

    _enrich(work, row, client, unmatched, dead)

    assert work["authors"] == ["Excel Author"]
    assert work["publisher"] == "Parser Publisher"
    assert work["release_date"] == "1980-05-20"
    assert work["release_precision"] == "day"
    assert work["cover_url"] == "https://example.com/cover.jpg"


def test_enrich_partial_excel_no_url_uses_opensearch():
    # No info_url in Excel — opensearch returns a URL.
    row = _row(
        info_url=None,
        cover_url=None,
        author=None,
        publisher=None,
        release_date_str=None,
    )
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = ("https://example.com/wiki/found", "opensearch")
    client.fetch_html.return_value = None  # dead URL after resolve

    unmatched, dead = [], []
    _enrich(work, row, client, unmatched, dead)
    assert any("dead_url" in entry for entry in unmatched)


def test_enrich_excel_uncredited_alone_falls_back_to_parser(monkeypatch):
    # Excel author cell contains only "Uncredited" — treat as missing and let
    # the parser fill it instead.
    row = _row(
        info_url="https://example.com/wiki/page",
        cover_url="https://example.com/cover.jpg",
        author="Uncredited",
        publisher="Some Pub",
        release_date_str="2020.05.01",
    )
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = (row.info_url, "from_excel")
    client.fetch_html.return_value = "<html>infobox</html>"

    monkeypatch.setattr(
        "scripts.build_data.parse_infobox",
        lambda html: {"authors": ["Real Author"]},
    )
    unmatched, dead = [], []
    _enrich(work, row, client, unmatched, dead)
    assert work["authors"] == ["Real Author"]


def test_enrich_excel_uncredited_mixed_keeps_real_names():
    row = _row(
        info_url="https://example.com/wiki/page",
        cover_url="https://example.com/cover.jpg",
        author="Real Author, Uncredited, Another",
        publisher="Some Pub",
        release_date_str="2020.05.01",
    )
    work = _row_to_work(row)
    client = MagicMock()
    client.resolve_url.return_value = (row.info_url, "from_excel")
    client.verify_url_alive.return_value = True

    unmatched, dead = [], []
    _enrich(work, row, client, unmatched, dead)
    assert work["authors"] == ["Real Author", "Another"]


def test_normalize_publisher_selects_us_edition():
    """Comma-separated publisher values get reduced to the US edition."""
    work = {"publisher": "Sphere Books (UK), Del Rey (US)"}
    build_data._normalize_publisher(work)
    assert work["publisher"] == "Del Rey"


def test_normalize_publisher_selects_worldwide_edition():
    work = {"publisher": "Dark Horse Comics (worldwide), Titan Magazines (UK)"}
    build_data._normalize_publisher(work)
    assert work["publisher"] == "Dark Horse Comics"


def test_normalize_publisher_selects_first_when_no_region_markers():
    work = {"publisher": "National Public Radio, HighBridge Audio"}
    build_data._normalize_publisher(work)
    assert work["publisher"] == "National Public Radio"


# ---------------------------------------------------------------------------
# _split_series_and_number tests
# ---------------------------------------------------------------------------


def test_split_series_and_number_single_value():
    assert _split_series_and_number("X-Wing", "5") == (["X-Wing"], ["5"])


def test_split_series_and_number_multi_value():
    assert _split_series_and_number("X-Wing, Rebel Alliance", "5, 12") == (
        ["X-Wing", "Rebel Alliance"],
        ["5", "12"],
    )


def test_split_series_and_number_strips_whitespace():
    assert _split_series_and_number("  X-Wing  ,Rebel Alliance ", "  5 , 12") == (
        ["X-Wing", "Rebel Alliance"],
        ["5", "12"],
    )


def test_split_series_and_number_drops_empties_from_double_commas():
    assert _split_series_and_number("X-Wing,,Rebel", "5,,12") == (
        ["X-Wing", "Rebel"],
        ["5", "12"],
    )


def test_split_series_and_number_truncates_extra_numbers():
    assert _split_series_and_number("X-Wing", "5, 12, 99") == (["X-Wing"], ["5"])


def test_split_series_and_number_short_number_list_allowed():
    assert _split_series_and_number("X-Wing, Rebel Alliance", "5") == (
        ["X-Wing", "Rebel Alliance"],
        ["5"],
    )


def test_split_series_and_number_empty_number():
    assert _split_series_and_number("X-Wing", None) == (["X-Wing"], [])


def test_split_series_and_number_empty_series_keeps_number():
    """Numbers are kept when series is empty; title acts as the series."""
    assert _split_series_and_number(None, "5") == ([], ["5"])
    assert _split_series_and_number("", "5") == ([], ["5"])
    assert _split_series_and_number(None, "1, 2") == ([], ["1", "2"])


def test_split_series_and_number_both_empty():
    assert _split_series_and_number(None, None) == ([], [])


# ---------------------------------------------------------------------------
# _row_to_work array-shape tests
# ---------------------------------------------------------------------------


def test_row_to_work_emits_series_as_array():
    row = _row(series="Star Wars Adventures (comics)", number="1")
    work = _row_to_work(row)
    assert work["series"] == ["Star Wars Adventures (comics)"]
    assert work["number"] == ["1"]


def test_row_to_work_emits_multi_series_arrays():
    row = _row(series="X-Wing, Rebel Alliance", number="5, 12")
    work = _row_to_work(row)
    assert work["series"] == ["X-Wing", "Rebel Alliance"]
    assert work["number"] == ["5", "12"]


def test_row_to_work_omits_series_when_empty():
    row = _row(series=None, number=None)
    work = _row_to_work(row)
    assert "series" not in work
    assert "number" not in work


def test_row_to_work_omits_number_when_empty_but_series_present():
    row = _row(series="The Clone Wars", number=None)
    work = _row_to_work(row)
    assert work["series"] == ["The Clone Wars"]
    assert "number" not in work


# ---------------------------------------------------------------------------
# derive_collection tests
# ---------------------------------------------------------------------------


def test_derive_collection_single_era_single_medium():
    from scripts.build_data import derive_collection
    from scripts.collections_reader import ExcelCollectionRow

    row = ExcelCollectionRow(
        title="Dark Empire (TPB)",
        release_date_str="1995.01.01",
        info_url=None,
        cover_url=None,
        color=None,
    )
    members = [
        {"id": "w-1", "era": "REBELLION", "medium": "Comic", "title": "Dark Empire 1", "year": 10},
        {"id": "w-2", "era": "REBELLION", "medium": "Comic", "title": "Dark Empire 2", "year": 11},
    ]
    c = derive_collection(row, members)
    assert c["title"] == "Dark Empire (TPB)"
    assert c["eras"] == ["REBELLION"]
    assert c["mediums"] == ["Comic"]
    assert c["year"] == 10
    assert c["year_end"] == 11
    assert c["anchor_year"] == 10
    assert c["anchor_era"] == "REBELLION"
    assert c["anchor_member_id"] == "w-1"
    assert c["member_ids"] == ["w-1", "w-2"]


def test_derive_collection_multi_era_multi_medium_dominant_medium():
    """Mediums sorted alphabetically. Dominant medium = Novel > Short Story.

    Range spans all members; anchor derived only from dominant members."""
    from scripts.build_data import derive_collection
    from scripts.collections_reader import ExcelCollectionRow

    row = ExcelCollectionRow(
        title="Tales of the Sith",
        release_date_str=None,
        info_url=None,
        cover_url=None,
        color=None,
    )
    members = [
        {
            "id": "w-ss-1",
            "era": "OLD REPUBLIC",
            "medium": "Short Story",
            "title": "SS 1",
            "year": -3000,
        },
        {
            "id": "w-novel",
            "era": "REBELLION",
            "medium": "Novel",
            "title": "Tales Novel",
            "year": 5,
            "year_end": 7,
        },
        {
            "id": "w-ss-2",
            "era": "NEW REPUBLIC",
            "medium": "Short Story",
            "title": "SS 2",
            "year": 20,
        },
    ]
    c = derive_collection(row, members)
    # Mediums + eras: union sorted.
    assert c["eras"] == ["NEW REPUBLIC", "OLD REPUBLIC", "REBELLION"]
    assert c["mediums"] == ["Novel", "Short Story"]
    # year/year_end: full range across ALL members.
    assert c["year"] == -3000
    assert c["year_end"] == 20
    # Anchor: dominant_medium = Novel, only one Novel member → use it.
    assert c["anchor_year"] == 5
    assert c["anchor_era"] == "REBELLION"
    assert c["anchor_member_id"] == "w-novel"


def test_derive_collection_omits_year_end_when_single_year():
    from scripts.build_data import derive_collection
    from scripts.collections_reader import ExcelCollectionRow

    row = ExcelCollectionRow(
        title="X", release_date_str=None, info_url=None, cover_url=None, color=None
    )
    members = [
        {"id": "w-1", "era": "REBELLION", "medium": "Comic", "title": "A", "year": 4},
        {"id": "w-2", "era": "REBELLION", "medium": "Comic", "title": "B", "year": 4},
    ]
    c = derive_collection(row, members)
    assert c["year"] == 4
    assert "year_end" not in c


def test_derive_collection_release_date_from_excel():
    from scripts.build_data import derive_collection
    from scripts.collections_reader import ExcelCollectionRow

    row = ExcelCollectionRow(
        title="X", release_date_str="1995.01.01", info_url=None, cover_url=None, color=None
    )
    members = [
        {"id": "w-1", "era": "REBELLION", "medium": "Comic", "title": "A", "year": 1},
        {"id": "w-2", "era": "REBELLION", "medium": "Comic", "title": "B", "year": 1},
    ]
    c = derive_collection(row, members)
    assert c["release_date"] == "1995-01-01"
    assert c["release_precision"] == "day"


# ---------------------------------------------------------------------------
# build() integration tests — collections wiring
# ---------------------------------------------------------------------------


def test_build_emits_collections_and_links_works(tmp_path, monkeypatch):
    """End-to-end: workbook with two members in a collection -> JSON has both
    arrays and works link via collection_ids."""
    from openpyxl import Workbook
    from scripts.build_data import build, EXCEL_PATH, OUTPUT_PATH
    import scripts.build_data as bd

    # Build a minimal workbook.
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("REBELLION")
    ws.append(["YEAR", "MEDIUM", "SERIES", "TITLE", "#", "AUTHOR",
               "PUBLISHER", "RELEASE", "COLLECTED", "INFO", "COVER"])
    ws.append(["10 ABY", "Comic", "Dark Empire", "Issue 1", 1, None, None, None,
               "Dark Empire (TPB)", None, None])
    ws.append(["10 ABY", "Comic", "Dark Empire", "Issue 2", 2, None, None, None,
               "Dark Empire (TPB)", None, None])
    cs = wb.create_sheet("COLLECTIONS")
    cs.append(["TITLE", "RELEASE", "INFO", "COVER"])
    cs.append(["Dark Empire (TPB)", "1995.01.01", None, None])
    wb_path = tmp_path / "x.xlsx"
    wb.save(wb_path)

    # Point build at the temp workbook and skip enrichment (dry_run=True
    # skips _enrich and avoids the network).
    monkeypatch.setattr(bd, "EXCEL_PATH", wb_path)
    payload = build(refresh=False, dry_run=True)

    assert "collections" in payload
    assert len(payload["collections"]) == 1
    c = payload["collections"][0]
    assert c["title"] == "Dark Empire (TPB)"
    assert c["eras"] == ["REBELLION"]
    assert c["mediums"] == ["Comic"]
    assert len(c["member_ids"]) == 2
    for w in payload["works"]:
        assert w["collection_ids"] == [c["id"]]


def test_build_handles_comma_separated_collected(tmp_path, monkeypatch):
    """A work with two collection titles ends up in both, in cell order."""
    from openpyxl import Workbook
    from scripts.build_data import build
    import scripts.build_data as bd

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("REBELLION")
    ws.append(["YEAR", "MEDIUM", "SERIES", "TITLE", "#", "AUTHOR",
               "PUBLISHER", "RELEASE", "COLLECTED", "INFO", "COVER"])
    ws.append(["5 ABY", "Short Story", "S", "Story 1", 1, None, None, None,
               "A (TPB), B (TPB)", None, None])
    ws.append(["5 ABY", "Short Story", "S", "Story 2", 2, None, None, None,
               "A (TPB)", None, None])
    ws.append(["5 ABY", "Short Story", "S", "Story 3", 3, None, None, None,
               "B (TPB)", None, None])
    cs = wb.create_sheet("COLLECTIONS")
    cs.append(["TITLE", "RELEASE", "INFO", "COVER"])
    cs.append(["A (TPB)", None, None, None])
    cs.append(["B (TPB)", None, None, None])
    wb_path = tmp_path / "x.xlsx"
    wb.save(wb_path)

    monkeypatch.setattr(bd, "EXCEL_PATH", wb_path)
    payload = build(refresh=False, dry_run=True)

    by_title = {c["title"]: c for c in payload["collections"]}
    assert set(by_title) == {"A (TPB)", "B (TPB)"}
    assert len(by_title["A (TPB)"]["member_ids"]) == 2
    assert len(by_title["B (TPB)"]["member_ids"]) == 2

    works_by_title = {w["title"]: w for w in payload["works"]}
    s1_ids = works_by_title["Story 1"]["collection_ids"]
    assert len(s1_ids) == 2
    assert s1_ids == [by_title["A (TPB)"]["id"], by_title["B (TPB)"]["id"]]
    assert works_by_title["Story 2"]["collection_ids"] == [by_title["A (TPB)"]["id"]]
    assert works_by_title["Story 3"]["collection_ids"] == [by_title["B (TPB)"]["id"]]


def test_build_partial_unmatched_keeps_valid_ids(tmp_path, monkeypatch):
    """`COLLECTED = "Valid, Bogus"` -> work keeps Valid's id, Bogus is logged."""
    from openpyxl import Workbook
    from scripts.build_data import build
    import scripts.build_data as bd

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("REBELLION")
    ws.append(["YEAR", "MEDIUM", "SERIES", "TITLE", "#", "AUTHOR",
               "PUBLISHER", "RELEASE", "COLLECTED", "INFO", "COVER"])
    ws.append(["5 ABY", "Comic", "S", "Issue 1", 1, None, None, None,
               "Valid (TPB), Bogus (TPB)", None, None])
    ws.append(["5 ABY", "Comic", "S", "Issue 2", 2, None, None, None,
               "Valid (TPB)", None, None])
    cs = wb.create_sheet("COLLECTIONS")
    cs.append(["TITLE", "RELEASE", "INFO", "COVER"])
    cs.append(["Valid (TPB)", None, None, None])
    wb_path = tmp_path / "x.xlsx"
    wb.save(wb_path)

    monkeypatch.setattr(bd, "EXCEL_PATH", wb_path)
    payload = build(refresh=False, dry_run=True)

    valid = payload["collections"][0]
    works_by_title = {w["title"]: w for w in payload["works"]}
    assert works_by_title["Issue 1"]["collection_ids"] == [valid["id"]]
    assert works_by_title["Issue 2"]["collection_ids"] == [valid["id"]]


def test_build_drops_single_member_collection(tmp_path, monkeypatch):
    from openpyxl import Workbook
    from scripts.build_data import build
    import scripts.build_data as bd

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("REBELLION")
    ws.append(["YEAR", "MEDIUM", "SERIES", "TITLE", "#", "AUTHOR",
               "PUBLISHER", "RELEASE", "COLLECTED", "INFO", "COVER"])
    ws.append(["10 ABY", "Comic", "Dark Empire", "Issue 1", 1, None, None, None,
               "Lonely (TPB)", None, None])
    cs = wb.create_sheet("COLLECTIONS")
    cs.append(["TITLE", "RELEASE", "INFO", "COVER"])
    cs.append(["Lonely (TPB)", None, None, None])
    wb_path = tmp_path / "x.xlsx"
    wb.save(wb_path)

    monkeypatch.setattr(bd, "EXCEL_PATH", wb_path)
    payload = build(refresh=False, dry_run=True)

    assert payload["collections"] == []
    assert len(payload["works"]) == 1
    assert "collection_ids" not in payload["works"][0]


def test_enrich_collection_populates_wiki_url_and_cover(tmp_path):
    """When INFO is blank, resolve by title; parse infobox for release+cover."""
    from scripts.build_data import _enrich_collection
    from scripts.collections_reader import ExcelCollectionRow

    class FakeClient:
        def resolve_url(self, info_url, title, series=None):
            assert info_url is None
            assert title == "Dark Empire (TPB)"
            return ("https://example/wiki/DE_TPB", "from_title")
        def fetch_html(self, url):
            return "<html><table class='infobox'>…</table></html>"
        def verify_url_alive(self, url):
            return True

    def fake_parser(html):
        return {
            "cover_url": "https://example/cover.jpg",
            "release_date": "1995-01-01",
            "release_precision": "day",
        }

    crow = ExcelCollectionRow(
        title="Dark Empire (TPB)", release_date_str=None,
        info_url=None, cover_url=None, color=None,
    )
    collection: dict = {"title": "Dark Empire (TPB)", "id": "c-1"}
    unmatched: list[str] = []
    _enrich_collection(
        collection, crow, FakeClient(), unmatched, parse_infobox=fake_parser,
    )
    assert collection["wiki_url"] == "https://example/wiki/DE_TPB"
    assert collection["cover_url"] == "https://example/cover.jpg"
    assert collection["release_date"] == "1995-01-01"
    assert collection["release_precision"] == "day"
    assert unmatched == []


def test_build_unmatched_collected_value_leaves_work_uncollected(
    tmp_path, monkeypatch,
):
    from openpyxl import Workbook
    from scripts.build_data import build
    import scripts.build_data as bd

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("REBELLION")
    ws.append(["YEAR", "MEDIUM", "SERIES", "TITLE", "#", "AUTHOR",
               "PUBLISHER", "RELEASE", "COLLECTED", "INFO", "COVER"])
    ws.append(["10 ABY", "Comic", "S", "Issue 1", 1, None, None, None,
               "Does Not Exist (TPB)", None, None])
    ws.append(["10 ABY", "Comic", "S", "Issue 2", 2, None, None, None, None,
               None, None])
    wb_path = tmp_path / "x.xlsx"
    wb.save(wb_path)

    monkeypatch.setattr(bd, "EXCEL_PATH", wb_path)
    payload = build(refresh=False, dry_run=True)
    assert payload["collections"] == []
    for w in payload["works"]:
        assert "collection_ids" not in w
