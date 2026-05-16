"""Read the COLLECTIONS sheet of Star Wars EU.xlsx into normalized rows."""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "COLLECTIONS"


@dataclass(frozen=True)
class ExcelCollectionRow:
    title: str
    release_date_str: str | None
    info_url: str | None
    cover_url: str | None
    color: str | None


def _stringify(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    return s or None


def read_collections(path: Path) -> Iterator[ExcelCollectionRow]:
    """Yield ExcelCollectionRow per non-empty data row in the COLLECTIONS sheet.

    Returns nothing if the sheet is absent (legacy workbook).
    """
    from scripts.excel_colors import ColorResolver

    wb = load_workbook(path, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return
        resolver = ColorResolver(wb)
        ws = wb[SHEET_NAME]
        for raw in ws.iter_rows(min_row=2):
            title = _stringify(raw[0].value)
            if not title:
                continue
            release_date_str = _stringify(raw[1].value) if len(raw) > 1 else None
            info_url = _stringify(raw[2].value) if len(raw) > 2 else None
            cover_url = _stringify(raw[3].value) if len(raw) > 3 else None
            color = resolver.resolve(raw[0])
            yield ExcelCollectionRow(
                title=title,
                release_date_str=release_date_str,
                info_url=info_url,
                cover_url=cover_url,
                color=color,
            )
    finally:
        wb.close()
