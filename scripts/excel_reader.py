"""Read Star Wars EU.xlsx into normalized rows."""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from scripts.year_utils import parse_year_range

# Sheet name -> era index. Order matches the spec's ERAS constant.
ERA_INDEX: dict[str, int] = {
    "PRE-REPUBLIC": 0,
    "OLD REPUBLIC": 1,
    "RISE OF THE EMPIRE": 2,
    "THE CLONE WARS": 3,
    "THE DARK TIMES": 4,
    "REBELLION": 5,
    "NEW REPUBLIC": 6,
    "NEW JEDI ORDER": 7,
    "LEGACY": 8,
    "NON-CANON": 9,
}

# Canonical medium spellings. Excel sometimes uses uppercase; normalize.
_MEDIUM_NORMALIZE: dict[str, str] = {
    "novel": "Novel",
    "junior novel": "Junior Novel",
    "young reader book": "Young Reader Book",
    "comic": "Comic",
    "short story": "Short Story",
    "movie": "Movie",
    "tv show": "TV Show",
    "videogame": "Videogame",
    "audio drama": "Audio Drama",
    "audio drama and comic": "Audio Drama",
    "audio drama and young reader book": "Audio Drama",
}


@dataclass(frozen=True)
class ExcelRow:
    era: int
    title: str
    series: str | None
    medium: str  # canonical Title Case (e.g. "Novel"); emitted as-is in works.json
    number: str | None
    year: int | None  # start year; may be None when Excel YEAR is empty
    year_end: int | None  # end year of a range; None for single-year entries
    info_url: str | None
    cover_url: str | None  # raw — may be ignored later in favor of wiki-fetched cover
    # Excel-as-source-of-truth fields. None when the cell is empty.
    author: str | None
    publisher: str | None
    release_date_str: str | None  # raw Excel "YYYY[.MM[.DD]]" string
    work_id: str | None = None  # canonical id from the ID column; None when blank


def _normalize_medium(raw: str | None) -> str:
    if not raw:
        return ""
    cleaned = " ".join(str(raw).strip().split())
    return _MEDIUM_NORMALIZE.get(cleaned.lower(), cleaned)


def _stringify(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    return s or None


def read_works(path: Path) -> Iterator[ExcelRow]:
    """Yield ExcelRow per non-empty data row in every sheet, in workbook order."""
    wb = load_workbook(path, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name not in ERA_INDEX:
                continue
            era = ERA_INDEX[sheet_name]
            ws = wb[sheet_name]
            for raw in ws.iter_rows(min_row=2):
                # Header layout: YEAR, MEDIUM, SERIES, TITLE, #, AUTHOR, PUBLISHER,
                # RELEASE, INFO, COVER, ID. Columns are read by absolute index.
                year_raw = _stringify(raw[0].value)
                medium_raw = _stringify(raw[1].value)
                series = _stringify(raw[2].value)
                title = _stringify(raw[3].value)
                number = _stringify(raw[4].value)
                author = _stringify(raw[5].value) if len(raw) > 5 else None
                publisher = _stringify(raw[6].value) if len(raw) > 6 else None
                release_date_str = _stringify(raw[7].value) if len(raw) > 7 else None
                info_url = _stringify(raw[8].value) if len(raw) > 8 else None
                cover_url = _stringify(raw[9].value) if len(raw) > 9 else None
                work_id = _stringify(raw[10].value) if len(raw) > 10 else None
                if not title or not medium_raw:
                    continue
                parsed = parse_year_range(year_raw)
                if parsed is None:
                    year_val: int | None = None
                    year_end_val: int | None = None
                else:
                    start, end = parsed
                    year_val = start
                    year_end_val = end if end != start else None
                yield ExcelRow(
                    era=era,
                    title=title,
                    series=series,
                    medium=_normalize_medium(medium_raw),
                    number=number,
                    year=year_val,
                    year_end=year_end_val,
                    info_url=info_url,
                    cover_url=cover_url,
                    author=author,
                    publisher=publisher,
                    release_date_str=release_date_str,
                    work_id=work_id,
                )
    finally:
        wb.close()
