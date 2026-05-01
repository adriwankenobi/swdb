"""Write enriched Wookieepedia fields back into the source Excel file."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from scripts.excel_reader import ERA_INDEX, _normalize_medium, _stringify

# Column indices (1-based) into the spreadsheet.
COL_AUTHOR = 6     # F
COL_PUBLISHER = 7  # G
COL_RELEASE = 8    # H
COL_COVER = 11     # K


def _make_lookup_key(
    era: int,
    title: str | None,
    series: str | None,
    medium: str,
    number: str | None,
) -> tuple:
    return (era, title, series, medium, number)


def _format_authors(authors: list[str]) -> str:
    return ", ".join(a.strip() for a in authors if a and a.strip())


def _format_release(iso_date: str, precision: str = "day") -> str:
    # "1976-11-12" -> "1976.11.12" (day), "1996.11" (month), "1996" (year)
    parts = iso_date.split("-")  # ["1976", "11", "12"]
    if precision == "year":
        return parts[0]
    if precision == "month":
        return ".".join(parts[:2])
    return ".".join(parts)


def update_excel(path: Path, enriched: dict[tuple, dict]) -> dict:
    """Write enriched fields back into the Excel file.

    `enriched` maps each lookup key (era, title, series, medium, number)
    to a dict possibly containing 'authors', 'publisher', 'release_date',
    'cover_url'. Trusted columns are never modified.

    Returns a summary dict with counts.
    """
    wb = load_workbook(path)
    try:
        updated = 0
        not_found = 0
        for sheet_name, era in ERA_INDEX.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2):
                medium_raw = _stringify(row[1].value)
                series = _stringify(row[2].value)
                title = _stringify(row[3].value)
                number = _stringify(row[4].value)
                if not title or not medium_raw:
                    continue
                medium = _normalize_medium(medium_raw)
                key = _make_lookup_key(era, title, series, medium, number)
                fields = enriched.get(key)
                if fields is None:
                    not_found += 1
                    continue
                changed = False
                if fields.get("authors") and not row[COL_AUTHOR - 1].value:
                    new_value = _format_authors(fields["authors"])
                    if new_value:
                        row[COL_AUTHOR - 1].value = new_value
                        changed = True
                if fields.get("publisher") and not row[COL_PUBLISHER - 1].value:
                    row[COL_PUBLISHER - 1].value = fields["publisher"]
                    changed = True
                if fields.get("release_date") and not row[COL_RELEASE - 1].value:
                    formatted = _format_release(
                        fields["release_date"],
                        fields.get("release_precision", "day"),
                    )
                    row[COL_RELEASE - 1].value = formatted
                    changed = True
                if fields.get("cover_url") and not row[COL_COVER - 1].value:
                    row[COL_COVER - 1].value = fields["cover_url"]
                    changed = True
                if changed:
                    updated += 1
        wb.save(path)
        return {"updated": updated, "not_found_in_excel": not_found}
    finally:
        wb.close()
